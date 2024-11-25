import pandas as pd
import numpy as np
from openai import OpenAI
import time
from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List
import random
import re
import pathlib as pl


@dataclass
class DataRow:
    slot: str
    value: str
    question: str
    answer: str
    slot_value_id: str
    turn_id: str
    text: str
    domain: str


class ExcelLoader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.data: List[DataRow] = []

    def load_data(self, sheet_name: str = None) -> List[DataRow]:
        """
        Load data from Excel file into a list of DataRow objects

        Args:
            sheet_name: Name of the sheet to load. If None, loads the active sheet.

        Returns:
            List of DataRow objects containing the Excel data
        """
        # Load the workbook
        wb = load_workbook(filename=self.file_path, read_only=True)

        # Get the specified sheet or active sheet
        ws = wb[sheet_name] if sheet_name else wb.active

        # Get headers from first row
        headers = [cell.value for cell in next(ws.rows)]

        # Process each row
        for row in ws.iter_rows(min_row=2):  # Skip header row
            # Get values from each cell in the row
            values = [str(cell.value) if cell.value is not None else "" for cell in row]

            # Create a dictionary of header-value pairs
            row_dict = dict(zip(headers, values))

            # Create DataRow object
            data_row = DataRow(
                slot=row_dict.get('slot', ''),
                value=row_dict.get('value', ''),
                question=row_dict.get('question', ''),
                answer=row_dict.get('answer', ''),
                slot_value_id=row_dict.get('slot_value_id', ''),
                turn_id=row_dict.get('turn_id', ''),
                text=row_dict.get('text', ''),
                domain=row_dict.get('domain', '').strip('"\'')
            )

            self.data.append(data_row)

        wb.close()
        return self.data

    def get_unique_domains(self) -> set:
        """Return set of unique domains in the data"""
        return {row.domain for row in self.data}

    def get_filtered_domains(self, exclude_domains: set) -> set:
        """Return set of domains excluding the ones in exclude_domains"""
        return {row.domain for row in self.data if row.domain not in exclude_domains}
    def get_filtered_data(self, exclude_domains: set) -> List[DataRow]:
        """Return data rows excluding specified domains"""
        return [row for row in self.data if row.domain not in exclude_domains]

    def filter_by_domain(self, domain: str) -> List[DataRow]:
        """Return rows filtered by domain"""
        return [row for row in self.data if row.domain == domain]

    def get_turn_data(self, turn_id: str) -> List[DataRow]:
        """Return all rows for a specific turn_id"""
        return [row for row in self.data if row.turn_id == turn_id]

    def print_domain_stats(self, exclude_domains: set = None):
        """Print statistics about domains"""
        all_domains = self.get_unique_domains()
        if exclude_domains:
            filtered_domains = self.get_filtered_domains(exclude_domains)
            print(f"Total domains: {len(all_domains)}")
            print(f"Excluded domains: {len(exclude_domains)}")
            print(f"Remaining domains: {len(filtered_domains)}")
            return filtered_domains
        return all_domains


client = OpenAI(
    api_key=pl.Path("~/.pw/openai.txt").expanduser().read_text().strip()
)



def generate_slot_names(qas, ans, text, q_a):
    slot_prompt = f"{q_a}Using the provided questions create a variable description to capture the underlying meaning of the question, and create a variable name that is general enough to be applied to multiple contexts. Keep each variable name short.Write each question before translating it into a value name, in the format: <question> : <variable description> \n <variable name>"
    response = client.chat.completions.create(
        model="gpt-4o-2024-05-13",
        messages=[
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": slot_prompt}
        ]
    )
    return response.choices[0].message.content

def generate_values(qas, ans, text, q_a):
    value_prompt = f"{q_a} Given the questions and their corresponding answers please create short value names that capture the information presented in the answer.  Write each answer before translating it into a value name, in the format: <Question> : <value name>"

    response = client.chat.completions.create(
        model="gpt-4o-2024-05-13",
        messages=[
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": value_prompt}
        ]
    )
    return response.choices[0].message.content

def process_domains(loader: ExcelLoader, exclude: ExcelLoader, num_domains: int = 100):
    """
    Process domains excluding the ones in exclude list and return specified number of unique domains
    """
    # Load the data
    data = loader.load_data()
    exclude_data = exclude.load_data()

    # Get excluded domains
    exclude_domains = exclude.get_unique_domains()

    # Get filtered domains and data
    filtered_domains = loader.get_filtered_domains(exclude_domains)
    filtered_data = loader.get_filtered_data(exclude_domains)

    # Print statistics
    #print("\nDomain Statistics:")
    #loader.print_domain_stats(exclude_domains)

    # Take only the specified number of domains
    random.seed(42)
    selected_domains = random.sample(list(filtered_domains), min(num_domains, len(filtered_domains)))
    #print(f"\nSelected {len(selected_domains)} domains for processing")

    # Get data for selected domains
    selected_data = []
    seen_domain = set()

    for i, row in enumerate(filtered_data):
        if row.domain in selected_domains and row.domain not in seen_domain:
            current_domain = row.domain
            seen_domain.add(current_domain)

            dialogue, j = [], i
            while filtered_data[j].domain == current_domain:
                dialogue.append(filtered_data[j])
                j += 1

            selected_data.append(dialogue)


    # print(f"Total rows in selected data: {len(selected_data)}")
    # for k, dialogue in enumerate(selected_data):
    #     print(f"Dialogue {k} (domain {dialogue[0].domain})has {len(dialogue)} rows")


    return selected_domains, selected_data


def extract_slots_and_descriptions(slot_names: str):
    """
    Extract slot descriptions and slot names from the API response using the updated regex.

    Args:
        slot_names: API response text containing slot descriptions and names.

    Returns:
        A tuple of two lists:
        - List of slot descriptions
        - List of slot names
    """
    # Updated regex to handle `:` as the delimiter
    pattern = r'(?P<question>.+?)\s*:\s*(?P<description>.+?)\s*:\s*(?P<variable_name>\w+)'
    descriptions = []
    slot_names_cleaned = []

    for match in re.finditer(pattern, slot_names):
        descriptions.append(match.group("description").strip())
        slot_names_cleaned.append(match.group("variable_name").strip())

    return descriptions, slot_names_cleaned


def get_first_response(response: str) -> str:
    """
    Extract the first response from a GPT output.

    Args:
        response: The complete GPT output containing multiple responses separated by '\n'.

    Returns:
        The first response string.
    """
    return response.split("\n")[0].strip()


#
#
# def extract_values(value_names: str) -> List[str]:
#     """
#     Extract value names from the API response.
#
#     Args:
#         value_names: API response text containing value names.
#
#     Returns:
#         A list of value names.
#     """
#     extracted_values = []
#
#     for entry in value_names.split("\n"):
#         entry = entry.strip()
#         if " : " in entry:
#             # Split into question and value
#             parts = entry.split(" : ", 1)
#             if len(parts) == 2:
#                 question, value = parts
#                 extracted_values.append(value.strip().strip('"'))
#
#     return extracted_values


if __name__ == "__main__":
    print(pl.Path("~/.pw/openai.txt").expanduser().read_text())

    # load the data and get all the rows.
    # Exclude the domain in the exclude file
    # Get 100 unique domains and their whole dialogue
    # send the dialogue in batches into gpt to get the regenerated versions ***
    # save the regenrated versions into a csv file


    # Initialize loaders
    loader = ExcelLoader('/Users/yasasvijosyula/Downloads/Summer_2024/updated_qa_with_domain.xlsx')
    exclude = ExcelLoader('/Users/yasasvijosyula/Downloads/Summer_2024/dot/excluded.xlsx')

    # Process domains and get filtered data
    selected_domains, selected_data = process_domains(loader, exclude)

    question_list, answer_list, text_list, slot_list = [], [], [], []

    master_slot_names_cleaned = []
    master_descriptions = []
    master_extracted_values = []

    all_results = []
    dialogue_results = {}

    for k, dialogue in enumerate(selected_data):
        counter = 0
        q_a_pairs = []
        for row in dialogue:
            question_list.append(row.question)
            answer_list.append(row.answer)
            text_list.append(row.text)
            slot_list.append(row.slot)


        for i in range(counter, counter + len(dialogue)):
            q_a_pairs.append("\n".join([f"{question_list[i]} -> {answer_list[i]}"]))
            counter += 1

        #print(q_a_pairs)
        #print(counter)

        for q_a in q_a_pairs:
            slot_names = generate_slot_names(question_list, answer_list, text_list, q_a)
            value_names = generate_values(question_list, answer_list, text_list, q_a)

            # Take only the first response
            slot_names = get_first_response(slot_names)
            value_names = get_first_response(value_names)

            print(f"QA: {q_a}\n")
            print(f"Slot: {slot_names}\n")
            print(f"Value: {value_names}\n")

            descriptions, slot_names_cleaned = extract_slots_and_descriptions(slot_names)

            master_descriptions.extend(descriptions)
            master_slot_names_cleaned.extend(slot_names_cleaned)

            # Process value names similarly
            value_names_list = value_names.split("\n")
            extracted_values = []
            for value_entry in value_names_list:
                value_entry = value_entry.strip()
                if " : " in value_entry:
                    parts = value_entry.split(" : ", 1)
                    if len(parts) == 2:
                        value = parts[1].strip().strip('"')
                        extracted_values.append(value)

            master_extracted_values.extend(extracted_values)

        print(len(q_a_pairs))
        print(len(text_list))
        print(len(master_slot_names_cleaned))
        #print(len(master_descriptions))
        print(len(master_extracted_values))
        #print(len(question_list))
        #print(len(answer_list))

        #print(master_slot_names_cleaned)

        print(master_slot_names_cleaned)
        print(master_extracted_values)

        print("NEW DIALOGUE")
        print(f"-"*80)

    #     dialogue_result_df = pd.DataFrame({
    #         'Text': text_list,
    #         'Slot': master_slot_names_cleaned,
    #         'Slot Description': master_descriptions,
    #         'Value': master_extracted_values,
    #         'Question': question_list,
    #         'Answer': answer_list
    #     })
    #
    #     all_results.append(dialogue_result_df)
    #     dialogue_results[f"Dialogue_{k + 1}"] = dialogue_result_df
    #
    # new_file_path = '/Users/yasasvijosyula/Downloads/Summer_2024/all_dialogues_results.xlsx'
    # with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
    #     for sheet_name, df in dialogue_results.items():
    #         df.to_excel(writer, sheet_name=sheet_name, index=False)
    #
    # print(f"All results saved to {new_file_path}")


    # Print some sample data for verification
    # print("\nSample of selected domains:")
    # for domain in list(selected_domains)[:5]:
    #     print(f"- {domain}")
    #     domain_data = loader.filter_by_domain(domain)
    #     print(f"  Number of rows: {len(domain_data)}")
    #
    # print("\nProcessing complete!")